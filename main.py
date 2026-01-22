import os
import time
import openpyxl
import random
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from helpers import extract_gpa, check_semester_exists

def main():
    # 1. Cấu hình trình duyệt
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    # Giả lập trình duyệt thật để tránh bị hệ thống quét IP nghi ngờ
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    
    # CHỐNG TREO: Đặt giới hạn load trang 20 giây và chờ phần tử 5 giây
    driver.set_page_load_timeout(20) 
    driver.implicitly_wait(5)
    
    excel_path = "Data_14DH.xlsx"
    BATCH_LIMIT = 100  
    processed_count = 0
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["14DHTH"]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if processed_count >= BATCH_LIMIT:
                print(f"🎯 Đạt hạn mức {BATCH_LIMIT} người. Dừng đợt này.")
                break
                
            status_hien_tai = row[7].value 
            url_xem_diem = row[4].value  

            if status_hien_tai and str(status_hien_tai).strip():
                continue

            if not url_xem_diem:
                continue

            try:
                # Dấu -u ở file .yml sẽ giúp dòng này hiện ra ngay lập tức
                print(f"🔍 [{processed_count + 1}/{BATCH_LIMIT}] Đang quét dòng {row_idx}...", flush=True)
                
                driver.get(str(url_xem_diem).strip())
                time.sleep(random.uniform(3, 5)) # Nghỉ lâu hơn một chút để an toàn
                
                soup = BeautifulSoup(driver.page_source, "html.parser")
                gpa = extract_gpa(soup)
                
                is_active = check_semester_exists(soup, "HK2 (2025 - 2026)")
                status_moi = "còn học" if is_active else "nghỉ học"
                
                ws.cell(row=row_idx, column=7, value=gpa)
                ws.cell(row=row_idx, column=8, value=status_moi)
                
                processed_count += 1
                
                # Lưu file ngay lập tức
                wb.save(excel_path)
                print(f"✅ Dòng {row_idx}: {gpa} | {status_moi}", flush=True)

            except Exception as e:
                print(f"⚠️ Bỏ qua dòng {row_idx} do web lag hoặc lỗi: {e}", flush=True)
                continue

        print(f"🏁 Xong đợt này. Quét được {processed_count} người.")

    finally:
        driver.quit()

if __name__ == "__main__":
    main()
