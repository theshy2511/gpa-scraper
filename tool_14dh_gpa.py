#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Tool thu thập GPA và trạng thái học tập cho sinh viên 14DH
Đọc URL từ cột E, lấy GPA thang 4, kiểm tra HK2 2025-2026
"""

import os
import time
import logging
import openpyxl
from bs4 import BeautifulSoup
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

# Import từ modules
from config import *
from helpers import extract_gpa, check_semester_exists

# ==== CONFIG CHO 14DH GPA ====
EXCEL_FILE_14DH = BASE_DIR / "Data_14DH.xlsx"
SHEET_14DH = "14DHTH"
LOG_FILE_14DH_GPA = LOGS_DIR / "tool_14dh_gpa.log"

# Target semester to check
TARGET_SEMESTER_14DH = "HK2 (2025 - 2026)"

# ==== SETUP LOGGING ====
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE_14DH_GPA, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def main():
    logger.info("="*60)
    logger.info("   BẮT ĐẦU THU THẬP GPA CHO 14DH")
    logger.info("="*60)
    
    if TEST_MODE:
        logger.info(f"⚠️  TEST MODE: Chỉ chạy {TEST_LIMIT} sinh viên đầu")
    
    # ==== KHỞI TẠO TRÌNH DUYỆT ====
    logger.info("🌐 Khởi tạo trình duyệt...")
    options = webdriver.ChromeOptions()
    if HEADLESS:
        options.add_argument("--headless")
    options.add_argument("--start-maximized")
    
    chromedriver_path = r"C:\chromedriver\chromedriver.exe"
    
    if os.path.exists(chromedriver_path):
        driver = webdriver.Chrome(service=Service(chromedriver_path), options=options)
    else:
        logger.warning("⚠️ ChromeDriver local không tìm thấy, thử download từ internet...")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    wait = WebDriverWait(driver, BROWSER_TIMEOUT)
    
    # ==== MỞ FILE EXCEL ====
    logger.info(f"📂 Mở file Excel: {EXCEL_FILE_14DH}")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE_14DH)
        ws = wb[SHEET_14DH]
    except Exception as e:
        logger.error(f"❌ Không thể mở Excel: {e}")
        driver.quit()
        return
    
    # ==== XỬ LÝ TỪNG SINH VIÊN ====
    sv_count = 0
    success_count = 0
    fail_count = 0
    skip_count = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        mssv_cell = row[0]  # Cột A: MSSV
        
        if not mssv_cell.value:
            continue
        
        mssv = str(mssv_cell.value).strip()
        ho_dem = row[1].value or ""
        ten = row[2].value or ""
        
        # ==== KIỂM TRA CỘT TRẠNG THÁI (H) ====
        # Chỉ xử lý nếu cột H rỗng HOẶC FORCE_REPROCESS = True
        status_cell = row[7] if len(row) > 7 else None  # Cột H (index 7)
        
        if not FORCE_REPROCESS:
            if status_cell and status_cell.value and str(status_cell.value).strip():
                logger.info(f"⏩ Bỏ qua dòng {row_idx} - {mssv} (đã có trạng thái: {status_cell.value})")
                skip_count += 1
                continue
        
        # Test mode: chỉ chạy N sinh viên
        if TEST_MODE and sv_count >= TEST_LIMIT:
            logger.info(f"⚠️  Đã đạt giới hạn TEST_LIMIT = {TEST_LIMIT}")
            break
        
        sv_count += 1
        
        logger.info("")
        logger.info(f"🔍 [{sv_count}] Đang xử lý: {mssv} - {ho_dem} {ten} (Dòng {row_idx})")
        
        # ==== ĐỌC URL TỪ CỘT E ====
        url_xem_diem = row[4].value if len(row) > 4 else None  # Cột E
        
        if not url_xem_diem or not str(url_xem_diem).strip():
            logger.warning(f"  ⚠️ Không có URL xem điểm - bỏ qua")
            fail_count += 1
            continue
        
        url_xem_diem = str(url_xem_diem).strip()
        logger.info(f"  🔗 URL: {url_xem_diem[:80]}...")
        
        try:
            # ==== TRUY CẬP TRANG ĐIỂM ====
            driver.get(url_xem_diem)
            time.sleep(2)  # Wait for page load
            
            # Parse bảng điểm
            soup = BeautifulSoup(driver.page_source, "html.parser")
            
            # ==== EXTRACT GPA ====
            gpa = extract_gpa(soup)
            logger.info(f"  📈 GPA: {gpa}")
            
            # ==== KIỂM TRA HK2 2025-2026 ====
            has_hk2 = check_semester_exists(soup, TARGET_SEMESTER_14DH)
            
            if has_hk2:
                status = "còn học"
                logger.info(f"  ✅ Trạng thái: {status}")
            else:
                status = "nghỉ học"
                logger.warning(f"  ⚠️ Trạng thái: {status}")
            
            # ==== GHI VÀO EXCEL ====
            # Cột G (7): GPA
            if gpa is not None:
                ws.cell(row=row_idx, column=7, value=gpa)
            
            # Cột H (8): Trạng thái
            ws.cell(row=row_idx, column=8, value=status)
            
            logger.info(f"  💾 Đã ghi: GPA={gpa}, Trạng thái={status}")
            success_count += 1
            
            # Lưu Excel sau mỗi sinh viên
            wb.save(EXCEL_FILE_14DH)
            logger.info(f"  💾 Đã lưu Excel")
            
            # Delay giữa các SV
            time.sleep(DELAY_BETWEEN_STUDENTS)
            
        except Exception as e:
            logger.error(f"❌ Lỗi xử lý MSSV {mssv}: {type(e).__name__}: {e}")
            fail_count += 1
            continue
    
    # ==== KẾT THÚC ====
    driver.quit()
    
    # Lưu Excel lần cuối
    try:
        wb.save(EXCEL_FILE_14DH)
        logger.info(f"💾 Đã lưu Excel: {EXCEL_FILE_14DH}")
    except Exception as e:
        logger.error(f"❌ Lỗi lưu Excel: {e}")
    
    wb.close()
    
    logger.info("")
    logger.info("="*60)
    logger.info("   KẾT THÚC THU THẬP GPA")
    logger.info("="*60)
    logger.info(f"📊 Tổng số SV cần xử lý: {sv_count}")
    logger.info(f"⏩ Đã bỏ qua (có sẵn):  {skip_count}")
    logger.info(f"✅ Thành công:          {success_count}")
    logger.info(f"❌ Thất bại:            {fail_count}")
    logger.info("")
    logger.info("🎉 HOÀN TẤT!")


if __name__ == "__main__":
    main()
