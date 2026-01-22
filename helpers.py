import io
import base64
import logging
import requests
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from config import CAPTCHA_API_URL, CAPTCHA_API_KEY

logger = logging.getLogger(__name__)

def solve_captcha_via_capsolver_from_bytes(image_bytes_io, timeout=30):
    """
    Gửi ảnh (io.BytesIO) tới CapSolver bằng createTask API.
    CAPTCHA format: 4 ký tự gồm CHỮ IN HOA + SỐ (KHÔNG phải 4 số)
    Trả về chuỗi captcha (solution.text) hoặc None nếu thất bại.
    """
    b64 = base64.b64encode(image_bytes_io.getvalue()).decode('utf-8')
    
    # Format ĐÚNG theo docs: https://docs.capsolver.com/en/guide/recognition/ImageToTextTask/
    payload = {
        "clientKey": CAPTCHA_API_KEY,
        "task": {
            "type": "ImageToTextTask",
            "websiteURL": "https://sinhvien.huit.edu.vn",  # Thêm để tăng accuracy
            "module": "common",  # Module: common (general OCR)
            "body": b64          # Base64 image
        }
    }
    
    headers = {"Content-Type": "application/json", "Accept": "application/json"}
    
    try:
        resp = requests.post(CAPTCHA_API_URL, json=payload, headers=headers, timeout=timeout)
        resp.raise_for_status()
    except requests.RequestException as e:
        logger.error(f"Lỗi gọi CAPTCHA API: {e}")
        return None

    try:
        j = resp.json()
    except ValueError:
        logger.error("CAPTCHA API trả về không phải JSON")
        return None

    # Kiểm tra lỗi API
    if j.get("errorId", 1) != 0:
        logger.error(f"CapSolver trả lỗi: {j.get('errorDescription') or j}")
        return None

    # Lấy solution.text
    sol = j.get("solution")
    if isinstance(sol, dict) and "text" in sol and isinstance(sol["text"], str):
        result = sol["text"].strip().upper()  # Force uppercase
        
        # Validation: chỉ check 4 ký tự (nới lỏng - chấp nhận cả toàn số)
        if len(result) == 4:
            logger.info(f"✅ CAPTCHA: {result}")
            return result
        else:
            logger.warning(f"⚠️ CAPTCHA không đúng 4 ký tự: {result} ({len(result)} ký tự)")
            return None
    
    logger.warning(f"Không tìm thấy 'solution.text' trong response: {j}")
    return None


def save_to_excel(subjects_data, output_path):
    """
    Lưu dữ liệu vào Excel format Power BI.
    subjects_data: List[Dict] - danh sách môn học
    """
    if not subjects_data:
        logger.warning("Không có dữ liệu để lưu")
        return
    
    # Tạo DataFrame
    df = pd.DataFrame(subjects_data)
    
    # Sắp xếp theo MSSV
    df = df.sort_values(by=['MSSV', 'TenMon'])
    
    # Lưu Excel
    df.to_excel(output_path, index=False, sheet_name="DuLieu", engine='openpyxl')
    
    logger.info(f"Đã lưu {len(subjects_data)} dòng dữ liệu")
    
    # Format columns (auto-fit)
    try:
        wb = load_workbook(output_path)
        ws = wb.active
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        logger.info(f"Đã format Excel: {output_path}")
    except Exception as e:
        logger.warning(f"Không thể format Excel: {e}")


def create_summary_report(subjects_data):
    """Tạo summary report từ dữ liệu thu thập được"""
    if not subjects_data:
        return "Không có dữ liệu"
    
    df = pd.DataFrame(subjects_data)
    
    total_students = df['MSSV'].nunique()
    total_subjects = len(df)
    avg_subjects_per_student = total_subjects / total_students if total_students > 0 else 0
    
    # Tính GPA trung bình
    student_gpas = []
    for mssv in df['MSSV'].unique():
        student_df = df[df['MSSV'] == mssv]
        total_tc = student_df['SoTC'].sum()
        total_diem = (student_df['DiemTK'] * student_df['SoTC']).sum()
        gpa = total_diem / total_tc if total_tc > 0 else 0
        student_gpas.append(gpa)
    
    avg_gpa = sum(student_gpas) / len(student_gpas) if student_gpas else 0
    
    report = f"""
═══════════════════════════════════════
         SUMMARY REPORT
═══════════════════════════════════════
📊 Tổng số sinh viên:      {total_students}
📚 Tổng số môn học:        {total_subjects}
📈 Trung bình môn/SV:      {avg_subjects_per_student:.1f}
🎯 GPA trung bình:         {avg_gpa:.2f}
═══════════════════════════════════════
    """
    return report


def save_checkpoint(checkpoint_data, checkpoint_file):
    """
    Lưu tiến trình xử lý vào file JSON để có thể resume sau này.
    checkpoint_data: dict với các key: last_processed_row, last_processed_mssv, timestamp, etc.
    """
    import json
    try:
        with open(checkpoint_file, 'w', encoding='utf-8') as f:
            json.dump(checkpoint_data, f, indent=2, ensure_ascii=False)
        logger.info(f"💾 Đã lưu checkpoint: {checkpoint_file}")
    except Exception as e:
        logger.warning(f"⚠️ Không thể lưu checkpoint: {e}")


def load_checkpoint(checkpoint_file):
    """
    Đọc tiến trình từ file JSON.
    Returns: dict hoặc None nếu file không tồn tại
    """
    import json
    import os
    
    if not os.path.exists(checkpoint_file):
        logger.info("📝 Không có checkpoint - bắt đầu từ đầu")
        return None
    
    try:
        with open(checkpoint_file, 'r', encoding='utf-8') as f:
            checkpoint = json.load(f)
        logger.info(f"📂 Đã load checkpoint: {checkpoint}")
        return checkpoint
    except Exception as e:
        logger.warning(f"⚠️ Không thể đọc checkpoint: {e}")
        return None


def extract_schedule_info(soup, subject_name):
    """
    Trích xuất thông tin giáo viên và mã học phần từ bảng lịch học.
    
    Args:
        soup: BeautifulSoup object của trang "Lịch theo tiến độ"
        subject_name: Tên môn học cần tìm (ví dụ: "Deep learning", "Lập trình di động")
    
    Returns:
        dict: {"teacher": "ThS. Đinh Thị Tâm", "course_code": "0101005281"}
        hoặc {"teacher": None, "course_code": None} nếu không tìm thấy
    """
    try:
        # Tìm bảng lịch học
        table = soup.find("table")
        if not table:
            logger.warning("⚠️ Không tìm thấy bảng lịch học")
            return {"teacher": None, "course_code": None}
        
        # Duyệt qua các dòng trong bảng
        rows = table.find_all("tr")
        
        for row in rows[1:]:  # Bỏ qua header
            cols = row.find_all("td")
            if len(cols) < 8:
                continue
            
            # Cột 2: Mã học phần, Cột 3: Tên môn học
            try:
                course_code = cols[1].text.strip()
                ten_mon = cols[2].text.strip()
                
                # Cột 8: Giảng viên
                teacher = cols[7].text.strip() if len(cols) > 7 else None
                
                # So khớp tên môn (case-insensitive, bỏ dấu)
                if subject_name.lower() in ten_mon.lower():
                    logger.info(f"  ✅ Tìm thấy: {ten_mon} | GV: {teacher} | Mã HP: {course_code}")
                    return {"teacher": teacher, "course_code": course_code}
            except Exception as e:
                logger.debug(f"  Skip row: {e}")
                continue
        
        logger.warning(f"  ⚠️ Không tìm thấy môn: {subject_name}")
        return {"teacher": None, "course_code": None}
        
    except Exception as e:
        logger.error(f"❌ Lỗi extract_schedule_info: {e}")
        return {"teacher": None, "course_code": None}


def write_to_13dh_excel(wb, sheet_name, row_index, subject_data):
    """
    Ghi dữ liệu điểm, giáo viên và mã học phần vào file Data_13DH.xlsx.
    
    Args:
        wb: openpyxl Workbook object
        sheet_name: Tên sheet (ví dụ: "13DHTH")
        row_index: Số thứ tự dòng trong Excel (1-indexed)
        subject_data: Dict chứa thông tin các môn học
            Format: {
                "Deep learning": {"grade": 3.5, "course_code": "01010051", "teacher": "TS. ABC"},
                "Lập trình di động": {"grade": 3.5, "course_code": "01010101", "teacher": "ThS. XYZ"},
                ...
            }
    
    Returns:
        bool: True nếu thành công, False nếu thất bại
    """
    try:
        ws = wb[sheet_name]
        
        # Mapping: Tên môn -> (cột điểm, cột giáo viên, cột mã học phần)
        subject_column_map = {
            "deep learning": (5, 6, 7),  # Cột E (điểm), F (GV), G (mã HP)
            "thực hành deep learning": (8, 9, 10),  # Cột H, I, J
            "lập trình di động": (11, 12, 13),  # Cột K, L, M
            "khai phá dữ liệu": (14, 15, 16),  # Cột N, O, P
            "quản trị hệ thống mạng": (17, 18, 19),  # Cột Q, R, S
            "thực hành quản trị hệ thống mạng": (20, 21, 22),  # Cột T, U, V
            "phân tích thiết kế hệ thống": (23, 24, 25),  # Cột W, X, Y
            "thực hành phân tích thiết kế hệ thống": (26, 27, 28)  # Cột Z, AA, AB
        }
        
        # Ghi dữ liệu vào Excel
        for subject_name, info in subject_data.items():
            subject_lower = subject_name.lower()
            
            if subject_lower in subject_column_map:
                grade_col, teacher_col, code_col = subject_column_map[subject_lower]
                
                # Ghi điểm
                if info.get("grade") is not None:
                    ws.cell(row=row_index, column=grade_col, value=info["grade"])
                
                # Ghi giáo viên
                if info.get("teacher"):
                    ws.cell(row=row_index, column=teacher_col, value=info["teacher"])
                
                # Ghi mã học phần
                if info.get("course_code"):
                    ws.cell(row=row_index, column=code_col, value=info["course_code"])
                
                logger.info(f"  📝 Đã ghi: {subject_name} -> Row {row_index}, Cols {grade_col}-{teacher_col}-{code_col}")
            else:
                logger.warning(f"  ⚠️ Không tìm thấy mapping cho môn: {subject_name}")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Lỗi write_to_13dh_excel: {e}")
        return False


# ==== V2 TOOL HELPERS ====

import re


def extract_gpa(soup):
    """
    Trích xuất GPA từ phần thống kê bên phải trang điểm.
    
    HTML Structure:
    <td class="text-start">Trung bình chung tích lũy: </td>
    <td><span>6,82</span> - <span>2,61</span> 30</td>
    
    Args:
        soup: BeautifulSoup object của trang "Xem điểm"
    
    Returns:
        float: GPA hoặc None nếu không tìm thấy
    """
    try:
        # Tìm tất cả <td> elements chứa text "Trung bình chung tích lũy"
        all_tds = soup.find_all('td')
        logger.info(f"  🔍 DEBUG: Total <td> elements found: {len(all_tds)}")
        
        # DEBUG: Show first 30 td texts to find the pattern
        for idx, td in enumerate(all_tds[:30]):
            td_text = td.get_text().strip()[:100]
            if 'trung' in td_text.lower() or 'bình' in td_text.lower() or 'tích' in td_text.lower():
                logger.info(f"  🔍 DEBUG td[{idx}]: '{td_text}'")
        
        for td in all_tds:
            td_text = td.get_text().strip()
            
            # Tìm td có text "Trung bình chung tích luỹ:" (chữ luỹ có dấu ỹ = y + móc)
            if re.search(r"Trung\s*bình\s*chung\s*tích\s*lu[ỹy]", td_text, re.IGNORECASE):
                logger.info(f"  🔍 Found GPA label: {td_text}")
                
                # Lấy td kế tiếp (next sibling)
                next_td = td.find_next_sibling('td')
                if next_td:
                    next_text = next_td.get_text()
                    logger.info(f"  🔍 Next td text: {next_text[:100]}")
                    
                    # Tìm số đầu tiên trong next_td (có thể trong <span> hoặc text node)
                    # Format: <span>6,82</span> - <span>2,61</span>
                    # Lấy số thứ 2 (GPA hệ 4) thay vì số thứ 1 (GPA hệ 10)
                    spans = next_td.find_all('span')
                    if len(spans) >= 2:
                        # Lấy số từ span thứ 2 (GPA hệ 4)
                        second_span_text = spans[1].get_text().strip()
                        match = re.search(r'(\d+[,\.]\d+)', second_span_text)
                        if match:
                            gpa_str = match.group(1).replace(',', '.')
                            gpa = float(gpa_str)
                            if 0 <= gpa <= 4:  # Validate GPA hệ 4 (0-4)
                                logger.info(f"  ✅ GPA extracted from span[1] (scale 4): {gpa}")
                                return gpa
                    
                    # Fallback: tìm tất cả số trong text và lấy số thứ 2
                    matches = re.findall(r'(\d+[,\.]\d+)', next_text)
                    if len(matches) >= 2:
                        gpa_str = matches[1].replace(',', '.')  # Số thứ 2
                        gpa = float(gpa_str)
                        if 0 <= gpa <= 4:
                            logger.info(f"  ✅ GPA extracted from text[1] (scale 4): {gpa}")
                            return gpa
        
        logger.warning(f"  ⚠️ Không tìm thấy GPA")
        return None
        
    except Exception as e:
        logger.error(f"  ❌ Lỗi extract_gpa: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return None


def extract_thesis_grade(soup):
    """
    Tìm và lấy điểm môn khóa luận tốt nghiệp.
    
    Args:
        soup: BeautifulSoup object của trang "Xem điểm"
    
    Returns:
        float: Điểm khóa luận hoặc None nếu không tìm thấy
    """
    try:
        table = soup.find("table", {"id": "xemDiem"})
        if not table:
            logger.warning("  ⚠️ Không tìm thấy bảng điểm")
            return None
        
        rows = table.find_all("tr")
        
        for tr in rows:
            cols = tr.find_all("td")
            if len(cols) < 8:
                continue
            
            try:
                # Cột 3: Tên môn
                ten_mon = cols[2].text.strip().lower()
                
                # Tìm môn khóa luận
                if "khóa luận" in ten_mon:
                    # Lấy điểm tổng kết
                    for td in cols:
                        if td.get("title") == "DiemTongKet":
                            diem_raw = td.text.strip().replace(",", ".")
                            try:
                                diem = float(diem_raw) if diem_raw else None
                                if diem is not None:
                                    logger.info(f"  ✅ Tìm thấy môn khóa luận: '{ten_mon}' | Điểm: {diem}")
                                    return diem
                            except:
                                pass
                            break
                    
            except Exception as e:
                continue
        
        logger.info("  ⏳ Không tìm thấy môn khóa luận")
        return None
        
    except Exception as e:
        logger.error(f"  ❌ Lỗi extract_thesis_grade: {e}")
        return None


def is_dropout_student(soup):
    """
    Kiểm tra xem sinh viên có nghỉ học không dựa trên:
    1. Học kỳ có quá nhiều điểm trống/0 (xét môn có TC > 0)
    2. Có nhiều môn "Không đạt" liên tục trong 1 kỳ
    
    Args:
        soup: BeautifulSoup object của trang "Xem điểm"
    
    Returns:
        tuple: (is_dropout: bool, thesis_grade: float or None)
            - is_dropout: True nếu nghỉ học, False nếu không
            - thesis_grade: Điểm khóa luận nếu có, None nếu không có
    """
    try:
        table = soup.find("table", {"id": "xemDiem"})
        if not table:
            logger.warning("  ⚠️ Không tìm thấy bảng điểm")
            return False, None
        
        rows = table.find_all("tr")
        
        # Group courses by semester
        semesters = {}
        current_semester = None
        thesis_grade = None
        
        for tr in rows:
            # Check for semester header (usually has colspan and contains "HK")
            if tr.find("td", {"colspan": True}):
                semester_text = tr.get_text().strip()
                if "HK" in semester_text or "Học kỳ" in semester_text:
                    current_semester = semester_text
                    semesters[current_semester] = []
                    logger.info(f"  📅 Tìm thấy học kỳ: {current_semester}")
                continue
            
            cols = tr.find_all("td")
            if len(cols) < 8:
                continue
            
            try:
                # Cột 2: Mã môn, Cột 3: Tên môn, Cột 4: Số TC
                ma_mon = cols[1].text.strip()
                ten_mon = cols[2].text.strip()
                ten_mon_lower = ten_mon.lower()
                
                # Skip header rows
                if ma_mon.lower() in ['mã môn', 'stt']:
                    continue
                
                # Get credits
                try:
                    so_tc = int(cols[3].text.strip())
                except:
                    so_tc = 0
                
                # Skip môn 0 TC (sinh hoạt, etc.)
                if so_tc == 0:
                    continue
                
                # Skip môn thể chất, quốc phòng
                if any(skip in ten_mon_lower for skip in ["thể chất", "quốc phòng", "sinh hoạt"]):
                    continue
                
                # Check for thesis grade
                if "khóa luận" in ten_mon_lower and thesis_grade is None:
                    for td in cols:
                        if td.get("title") == "DiemTongKet":
                            diem_raw = td.text.strip().replace(",", ".")
                            try:
                                thesis_grade = float(diem_raw) if diem_raw else None
                            except:
                                pass
                            break
                
                # Get grade and status
                diem_tk = None
                diem_chu = None
                
                for td in cols:
                    if td.get("title") == "DiemTongKet":
                        diem_raw = td.text.strip().replace(",", ".")
                        try:
                            diem_tk = float(diem_raw) if diem_raw else None
                        except:
                            diem_tk = None
                        break
                
                # Get điểm chữ (usually second to last or last column)
                if len(cols) >= 7:
                    diem_chu = cols[-3].text.strip() if len(cols) > 7 else cols[-2].text.strip()
                
                course_info = {
                    'ma_mon': ma_mon,
                    'ten_mon': ten_mon,
                    'so_tc': so_tc,
                    'diem_tk': diem_tk,
                    'diem_chu': diem_chu
                }
                
                if current_semester:
                    semesters[current_semester].append(course_info)
                
            except Exception as e:
                logger.debug(f"  Skip row: {e}")
                continue
        
        # Analyze semesters for dropout detection
        for semester_name, courses in semesters.items():
            if not courses:
                continue
            
            # Filter courses with TC > 0
            courses_with_credits = [c for c in courses if c['so_tc'] > 0]
            
            if not courses_with_credits:
                continue
            
            # Count missing grades
            missing_count = sum(1 for c in courses_with_credits if c['diem_tk'] is None or c['diem_tk'] == 0)
            
            # Count "Không đạt" (F, D, D+)
            failed_count = sum(1 for c in courses_with_credits if c.get('diem_chu') in ['F', 'D', 'D+'])
            
            total_count = len(courses_with_credits)
            
            # Dropout condition 1: > 50% môn thiếu điểm
            if missing_count / total_count > 0.5:
                logger.warning(f"  ⚠️ DROPOUT detected: {semester_name} - {missing_count}/{total_count} môn thiếu điểm")
                return True, thesis_grade
            
            # Dropout condition 2: > 50% môn không đạt
            if failed_count / total_count > 0.5:
                logger.warning(f"  ⚠️ DROPOUT detected: {semester_name} - {failed_count}/{total_count} môn không đạt")
                return True, thesis_grade
        
        # Not dropout
        return False, thesis_grade
        
    except Exception as e:
        logger.error(f"  ❌ Lỗi is_dropout_student: {e}")
        return False, None


def write_to_gpa_v2_excel(wb, sheet_name, row_index, data):
    """
    Ghi dữ liệu GPA và điểm khóa luận vào file Data_13DH_V2.xlsx.
    
    Args:
        wb: openpyxl Workbook object
        sheet_name: Tên sheet
        row_index: Số thứ tự dòng trong Excel (1-indexed)
        data: Dict chứa thông tin:
            {
                'link': str - URL trang xem điểm,
                'gpa': float - GPA,
                'thesis_grade': float - Điểm khóa luận (0 = nghỉ học, 36 = chưa đăng ký),
                'status': str - 'completed', 'dropout', 'not_registered'
            }
    
    Returns:
        bool: True nếu thành công, False nếu thất bại
    """
    try:
        ws = wb[sheet_name]
        
        # Column mapping:
        # E (5): Link
        # F (6): GPA
        # G (7): Điểm khóa luận
        # H (8): Trạng thái (optional)
        
        # Ghi link
        ws.cell(row=row_index, column=5, value=data.get('link'))
        
        # Ghi GPA
        if data.get('gpa') is not None:
            ws.cell(row=row_index, column=6, value=data['gpa'])
        
        # Ghi điểm khóa luận
        thesis = data.get('thesis_grade')
        if thesis is not None:
            ws.cell(row=row_index, column=7, value=thesis)
        
        # Ghi trạng thái (optional - for debugging)
        status_map = {
            'completed': 'Có điểm',
            'dropout': 'Nghỉ học',
            'not_registered': 'Chưa đăng ký'
        }
        status_text = status_map.get(data.get('status'), '')
        ws.cell(row=row_index, column=8, value=status_text)
        
        logger.info(f"  📝 Đã ghi: Row {row_index} | Link: {data.get('link')[:50]}... | GPA: {data.get('gpa')} | KL: {thesis} | Status: {status_text}")
        
        return True
        
    except Exception as e:
        logger.error(f"❌ Lỗi write_to_gpa_v2_excel: {e}")
        return False


def check_semester_exists(soup, semester_name):
    """
    Check if a specific semester exists in the grade table.
    
    Args:
        soup: BeautifulSoup object of grade page
        semester_name: String to match (e.g., "HK2 (2025 - 2026)")
    
    Returns:
        bool: True if semester found, False otherwise
    """
    try:
        table = soup.find("table", {"id": "xemDiem"})
        if not table:
            logger.warning("  ⚠️ Không tìm thấy bảng điểm")
            return False
        
        rows = table.find_all("tr")
        
        for tr in rows:
            # Check for semester header (usually has colspan and contains "HK")
            if tr.find("td", {"colspan": True}):
                semester_text = tr.get_text().strip()
                
                # Normalize both strings for comparison (remove extra spaces, case insensitive)
                semester_text_normalized = " ".join(semester_text.split()).lower()
                semester_name_normalized = " ".join(semester_name.split()).lower()
                
                # Check if semester name is in the text
                if semester_name_normalized in semester_text_normalized:
                    logger.info(f"  ✅ Tìm thấy học kỳ: {semester_text}")
                    return True
        
        logger.info(f"  ⏳ Không tìm thấy học kỳ: {semester_name}")
        return False
        
    except Exception as e:
        logger.error(f"  ❌ Lỗi check_semester_exists: {e}")
        return False
